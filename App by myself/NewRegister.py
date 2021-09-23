from PIL import ImageTk
from openpyxl import *
from tkinter import *
from tkinter import messagebox


def toplevel(root):
    global name_field, pass_field, email_field, address_field, born_field, phone_field, university_field
    # set toplevel
    re = Toplevel(root)
    re.title("Register")
    re.geometry("1100x600+300+0")
    re.resizable(0, 0)

    # set background
    img = ImageTk.PhotoImage(file="D:\App by myself\Images\Snature.jpg")
    background = Label(re, image= img, highlightthickness=0, borderwidth= 0)
    background.place(x= 0, y= 0)

    # set frame
    frame = Frame(re, background="#2d3134",height=325, width= 800)
    frame.place(x= 100, y= 80)

    # set button
    Submit = Button(frame, text= "Submit", background="#2d3134", font = ("", 15, 'bold'), fg= "#f0f0f0", command= submit)
    Delete = Button(frame, text= "Delete All",  background="#2d3134", font = ("", 15, 'bold'), fg= "#f0f0f0", command= clear)

    # set Label Quit
    Quit_label = Label(frame, text= "Quit app ?", font= (8), background= "#2d3134", fg= "red")

    #set command on quit label
    def Quit_app(event):
        re.destroy()
    Quit_label.bind("<Button-1>", Quit_app)
    # set focus In and Out Quit
    def enter_quit(event):
        Quit_label.configure( font= ("", 12, "underline"), fg= "white")
    def exit_quit(event):
        Quit_label.configure(  font= (9),  fg= "red")

    Quit_label.bind("<Enter>", enter_quit)
    Quit_label.bind("<Leave>", exit_quit)

    
    # create Label
    title_label = Label(frame, text= "Registration Form", background="#2d3134", font = ("", 15, 'bold'), fg= "#f0f0f0")
    name_label = Label(frame, text= "NameUser", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    pass_label = Label(frame, text= "PassWord", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    email_label =Label(frame, text= "Email", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    address_label = Label(frame, text= "Address", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    born_label =Label(frame, text= "Born", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    phone_label =Label(frame, text= "Phone Number", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    university_label = Label(frame, text= "University", background="#2d3134", font = ("", 10, 'bold'), fg= "#f0f0f0")
    # create Entry
    name_field = Entry(frame, width=80)
    pass_field = Entry(frame, width=80)
    email_field = Entry(frame, width=80)
    address_field = Entry(frame, width=80)
    born_field = Entry(frame, width=80)
    phone_field = Entry(frame, width=80)
    university_field = Entry(frame, width=80)
    # set position for widget
    title_label.place(x= 150, y= 6)
    name_label.place(x= 10, y= 40)
    pass_label.place(x= 10, y= 70)
    email_label.place(x= 10, y= 100)
    address_label.place(x= 10, y= 130)
    born_label.place(x= 10, y= 160)
    phone_label.place(x= 10, y= 190)
    university_label.place(x= 10, y= 220)

    name_field.place(x= 120, y= 40)
    pass_field.place(x= 120, y= 70)
    email_field.place(x= 120, y= 100)
    address_field.place(x= 120, y= 130)
    born_field.place(x= 120, y= 160)
    phone_field.place(x= 120, y= 190)
    university_field.place(x= 120, y= 220)

    Submit.place(x= 650, y= 130)
    Delete. place(x= 300, y= 270)

    Quit_label.place(x= 720, y= 295)

    # define funxtion focus field  when press Enter
    def pass_focus(event):
        pass_field.focus_set()
    def email_focus(event):
        email_field.focus_set()
    def address_focus(event):
        address_field.focus_set()
    def born_focus(event):
        born_field.focus_set()
    def phone_focus(event):
        phone_field.focus_set()
    def university_focus(event):
        university_field.focus_set()
    # set focus field 
    name_field.focus_set()
    name_field.bind("<Return>", pass_focus)
    pass_field.bind("<Return>", email_focus)
    email_field.bind("<Return>", address_focus)
    address_field.bind("<Return>", born_focus)
    born_field.bind("<Return>", phone_focus)
    phone_field.bind("<Return>", university_focus)

    mainloop()

def clear():
        name_field.delete(0, END)
        pass_field.delete(0, END)
        email_field.delete(0, END)
        address_field.delete(0, END)
        born_field.delete(0, END)
        phone_field.delete(0, END)
        university_field.delete(0, END)

# define function wwhen press Submit
def submit():
        if (name_field.get() == "" or
        pass_field.get() == "" or 
        email_field.get() == "" or 
        address_field.get() == "" or 
        born_field.get() == "" or 
        phone_field.get() == "" or 
        university_field.get() == ""):
            messagebox.showerror("Empty", "Please fill in all the information")

        else:
            wb = load_workbook("Excel\AccountUser.xlsx")
            sheet = wb.active
            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet.max_row

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet.cell(row= current_row + 1, column= 1).value = str(current_row)
            sheet.cell(row= current_row + 1, column= 2).value = name_field.get()
            sheet.cell(row= current_row + 1, column= 3).value = pass_field.get()
            sheet.cell(row= current_row + 1, column= 4).value = email_field.get()
            sheet.cell(row= current_row + 1, column= 5).value = address_field.get()
            sheet.cell(row= current_row + 1, column= 6).value = born_field.get()
            sheet.cell(row= current_row + 1, column= 7).value = phone_field.get()
            sheet.cell(row= current_row + 1, column= 8).value = university_field.get()
            # save file
            wb.save("Excel\AccountUser.xlsx")
            # clear all
            clear()
            # set focus on name_field
            name_field.focus_set()



