from tkinter import *
from PIL import ImageTk
import NewRegister as Re
from openpyxl import *
import TimeManagement as Tim
from tkinter import messagebox


"""+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"""

def Login(user,pas,root):
    # open file excel
    wb = load_workbook("Excel\AccountUser.xlsx")
    ws = wb.active
    # set max row on var
    rows= ws.max_row

    lst = {}
    # append user name in lst
    for i in range(2,rows+1):
        a = ws["B"+str(i)].value
        lst[a] = str(ws["C"+str(i)].value)

    # check empty 
    if ( user == "" or pas == ""):
        messagebox.showerror("Error", "Please fill in all the information")
    # check user name
    else:        
        for i in lst.keys():
            if user == i:
                # check passwword
                if pas == lst[i]:
                    Tim.toplevel(root)
                    return          
                else:
                    messagebox.showerror("Error", "Password's wrong")
                    return 
        messagebox.showerror("Error","Account don't exist")

"""+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"""



main = Tk()
main.title("AByMS app")
main.geometry("1050x490+300+180")
main.resizable(0, 0)

"""_________________________________________________________________________________________________________________________________________________________________________________"""
# Set background for GUI
count = -1
lst = [
    ImageTk.PhotoImage(file="Images\Space.jpg"),
    ImageTk.PhotoImage(file="Images\Space1.jpg"),
    ImageTk.PhotoImage(file= "Images\Space2.jpg"),
    ImageTk.PhotoImage(file= "Images\Space3.jpg"),
    ImageTk.PhotoImage(file= "Images\Space4.jpg"),
]

background = Label(main, image= lst[0], highlightthickness= 0, borderwidth= 0)
background.place(x= 0, y= 0)

def next():
    global count
    if count == 4:
        background.configure(image= lst[0])
        count= 0
    else:
        background.configure(image= lst[count + 1])
        count += 1
    main.after(1500, next)
# running 
next()    

"""_________________________________________________________________________________________________________________________________________________________________________________"""




"""+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"""

# Define Open_login
def Open_login(root):
    user = user_field.get()
    pas = pass_field.get()
    Login(user,pas,root)
    

# Define Open_register
def Open_register(event):
    Re.toplevel(main)

"""+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"""



"""_________________________________________________________________________________________________________________________________________________________________________________"""

# set frame contains widgets
frame = Frame(main, bg= "#2d3134", height= 490, width= 300)


# Set label
wel = Label(frame, text="Welcome to AByMS", font= ("", 15, "bold"), fg= "#f0f0f0", bg= "#2d3134")
login_lab = Label(frame, text= "Login", font= ("", 25 , "bold"), fg = "#f0f0f0", bg= "#2d3134")
user_text = Label(frame, text= "User name :", font= ("", 10, "bold"), fg= "#f0f0f0", bg= "#2d3134")
pass_text = Label(frame, text= "  Password :", font= ("", 10, "bold"), fg= "#f0f0f0", bg= "#2d3134")
Ds_label = Label(frame, text= "DSer K47 UEH :))))", font= ("", 10, "bold"), fg= "#f0f0f0", bg= "#2d3134")


# set Entry
user_field = Entry(frame, width= 30, bg= "gray", fg= "white", borderwidth= 0)
pass_field = Entry(frame, width= 30, bg= "gray", fg= "white", borderwidth= 0)


# set Button Login
Login_but = Button(frame, text= "Login",  relief= RAISED,
                                            font = ('', 12, 'bold'), foreground = 'light blue', 
                                            background= 'gray', borderwidth= 0, 
                                            width= 12, height= 1,
                                            command= lambda: Open_login(main))


# set Label Register 
Re_label = Label(frame, text= "Register ?", font= (5), background= "#2d3134", fg= "light green")
Re_label.bind("<Button-1>", Open_register)


# set Label Quit
Quit_label = Label(frame, text= "Quit app ?", font= (8), background= "#2d3134", fg= "red")

#set command on quit label
def Quit_app(event):
    main.destroy()
Quit_label.bind("<Button-1>", Quit_app)


# set position for widgets
frame.place(x= 750, y= 0)
wel.place(x= 0, y= 0)
login_lab.place(x= 10, y= 80)
user_text.place( x= 10, y= 150)
pass_text.place( x= 10, y= 183)
user_field.place(x= 100, y= 153)
pass_field.place(x= 100, y= 185)
Login_but.place(x= 70, y= 215)
Re_label.place(x= 220, y= 205 )
Quit_label.place(x= 220, y= 460)
Ds_label.place(x= 90, y= 350)

#set focus widget
def pass_focus(event):
    pass_field.focus_set()

user_field.focus_set()
user_field.bind("<Return>", pass_focus)

# set Focus In and Out Register
def enter_re(event):
    Re_label.configure( font= ("", 12, "underline"), fg= "white")
def exit_re(event):
    Re_label.configure(  font= (9),  fg= "light green")

Re_label.bind("<Enter>", enter_re)
Re_label.bind("<Leave>", exit_re)

# set focus In and Out Quit
def enter_quit(event):
    Quit_label.configure( font= ("", 12, "underline"), fg= "white")
def exit_quit(event):
    Quit_label.configure(  font= (9),  fg= "red")

Quit_label.bind("<Enter>", enter_quit)
Quit_label.bind("<Leave>", exit_quit)

mainloop()

"""_________________________________________________________________________________________________________________________________________________________________________________"""







