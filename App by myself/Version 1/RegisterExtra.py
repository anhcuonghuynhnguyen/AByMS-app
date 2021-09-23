from openpyxl import *
from tkinter import *
from tkinter import messagebox
def excel():
	global wb, sheet
	wb = load_workbook("Excel\AccountUser.xlsx")
	sheet = wb.active
    # resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 5
	sheet.column_dimensions['B'].width = 30
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 40
	sheet.column_dimensions['E'].width = 30
	sheet.column_dimensions['F'].width = 12
	sheet.column_dimensions['G'].width = 10

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "STT"
	sheet.cell(row=1, column=2).value = "Name"
	sheet.cell(row=1, column=3).value = "Password"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Address"
	sheet.cell(row=1, column=6).value = "Born"
	sheet.cell(row=1, column=7).value = "Phone Number"
	sheet.cell(row=1, column=8).value = "University"
	# save file
	wb.save("Excel\AccountUser.xlsx")

def toplevel(root):
	global name_field, pass_field, email_field, address_field, born_field, phone_field, university_field
    # tkinter
	re = Toplevel(root)
	re.title("Register")
	re.geometry("700x370+500+200")
	re.resizable(0, 0)

	frame = Frame(re, background="light blue",height=300, width= 580)
	frame.place(x= 60, y= 80)
	
	#create Label
	label = Label(re, text= "Form Register", font = ("Time New Roman", 20, 'bold'), fg= "blue")
	label.pack(side= "top", pady= 10)

	#create Button
	submit = Button(re, text= "Submit", font = ("Time New Roman", 20, 'bold'),fg="green", command= lambda: Submit())
	submit.pack(side= "bottom", pady= 10)

	escape = Button(re, text= "Escape", font = ("Time New Roman", 10, 'bold'), fg= "red", command= re.destroy)
	escape.place(x=10,y= 10)

	# create Label
	name_label = Label(frame, text= "NameUser", background="light blue", font = ("Time New Roman", 10, 'bold'))
	pass_label = Label(frame, text= "PassWord", background="light blue", font = ("Time New Roman", 10, 'bold'))
	email_label =Label(frame, text= "Email", background="light blue", font = ("Time New Roman", 10, 'bold'))
	address_label = Label(frame, text= "Address", background="light blue", font = ("Time New Roman", 10, 'bold'))
	born_label =Label(frame, text= "Born", background="light blue", font = ("Time New Roman", 10, 'bold'))
	phone_label =Label(frame, text= "Phone Number", background="light blue", font = ("Time New Roman", 10, 'bold'))
	university_label = Label(frame, text= "University", background="light blue", font = ("Time New Roman", 10, 'bold'))
	# create Entry
	name_field = Entry(frame, width=80)
	pass_field = Entry(frame, width=80)
	email_field = Entry(frame, width=80)
	address_field = Entry(frame, width=80)
	born_field = Entry(frame, width=80)
	phone_field = Entry(frame, width=80)
	university_field = Entry(frame, width=80)
	# set position for widget
	name_label.grid(row= 0, column= 0, padx=10, pady= 4)
	pass_label.grid(row= 1, column= 0, padx=10, pady= 4)
	email_label.grid(row= 2, column= 0, padx=10, pady= 4)
	address_label.grid(row= 3, column= 0, padx=10, pady= 4) 
	born_label.grid(row= 4, column= 0, padx=10, pady= 4) 
	phone_label.grid(row= 5, column= 0, padx=10, pady= 4) 
	university_label.grid(row= 6, column= 0, padx=10, pady= 4) 

	name_field.grid(row= 0, column= 1, padx=10, pady= 2) 
	pass_field.grid(row= 1, column= 1, padx=10, pady= 2) 
	email_field.grid(row= 2, column= 1, padx=10, pady= 2) 
	address_field.grid(row= 3, column= 1, padx=10, pady= 2) 
	born_field.grid(row= 4, column= 1, padx=10, pady= 2) 
	phone_field.grid(row= 5, column= 1, padx=10, pady= 2) 
	university_field.grid(row= 6, column= 1, padx=10, pady= 2) 
	
	# define funxtion focus field  when press Enter
	def pass_focus(event):
		pass_field.focus_set()
	def email_focus(event):
		email_field.focus_set()
	def address_focus(event):
		address_field.focus_set()
	def born_focus(event):
		born_field.focus_set()
	def phone_focus(event):
		phone_field.focus_set()
	def university_focus(event):
		university_field.focus_set()
	# set focus field 
	name_field.focus_set()
	name_field.bind("<Return>", pass_focus)
	pass_field.bind("<Return>", email_focus)
	email_field.bind("<Return>", address_focus)
	address_field.bind("<Return>", born_focus)
	born_field.bind("<Return>", phone_focus)
	phone_field.bind("<Return>", university_focus)

# define function wwhen press Submit
def Submit():
	if (name_field.get() == "" or
		pass_field.get() == "" or 
		email_field.get() == "" or 
		address_field.get() == "" or 
		born_field.get() == "" or 
		phone_field.get() == "" or 
		university_field.get() == ""):
		messagebox.showerror("Empty", "Please fill in all the information")
	
	else:
		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row= current_row + 1, column= 1).value = str(current_row)
		sheet.cell(row= current_row + 1, column= 2).value = name_field.get()
		sheet.cell(row= current_row + 1, column= 3).value = pass_field.get()
		sheet.cell(row= current_row + 1, column= 4).value = email_field.get()
		sheet.cell(row= current_row + 1, column= 5).value = address_field.get()
		sheet.cell(row= current_row + 1, column= 6).value = born_field.get()
		sheet.cell(row= current_row + 1, column= 7).value = pass_field.get()
		sheet.cell(row= current_row + 1, column= 8).value = university_field.get()
		# save file
		wb.save("Excel\AccountUser.xlsx")
		# clear all
		clear()
		# set focus on name_field
		name_field.focus_set()

def clear():
	name_field.delete(0, END)
	pass_field.delete(0, END)
	email_field.delete(0, END)
	address_field.delete(0, END)
	born_field.delete(0, END)
	phone_field.delete(0, END)
	university_field.delete(0, END)

